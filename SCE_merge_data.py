import pandas as pd
from openpyxl import load_workbook

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

# Paths to CSV and XLSX files
csv_file_path = "/path/to/sce_usage_parsed.csv"
xlsx_file_path = "/path/to/SCE Usage Pivot (MASTER).xlsx"

# Worksheet and table names
worksheet_name = "sce_usage_parsed"
table_name = "Table1"

try:
    # Read CSV data
    csv_data = pd.read_csv(csv_file_path)
    csv_data = csv_data.map(lambda x: x.strip() if isinstance(x, str) else x)
    if 'Date' in csv_data.columns:
        csv_data['Date'] = pd.to_datetime(csv_data['Date']).dt.date
    if 'StartTime' 'EndTime' in csv_data.columns:
        csv_data['StartTime', 'EndTime'] = pd.to_datetime(csv_data['StartTime', 'EndTime']).dt.time
        csv_data['StartTime', 'EndTime'] = csv_data['StartTime', 'EndTime'].apply(lambda x: x.strftime('%I:%M:%S %p'))
    if 'kwHUsage' 'Cost' in csv_data.columns:
        csv_data['kwHUsage' 'Cost'] = pd.to_numeric(csv_data['kwHUsage' 'Cost'], errors='coerce')
    print("Opened CSV file...")
except Exception as e:
    print(f"Error reading CSV file: {e}")

try:
    # Load existing XLSX file
    workbook = load_workbook(xlsx_file_path)

    # Get the specific worksheet
    worksheet = workbook[worksheet_name]

    # Identify number formats from existing data
    number_formats = {}
    for col_idx, column in enumerate(worksheet.iter_cols(min_col=1, max_col=len(csv_data.columns), min_row=1, max_row=1)):
        number_formats[col_idx] = column[0].number_format

    # Append CSV data to the bottom of the table
    for row in csv_data.itertuples(index=False):
        worksheet.append(row)

    # Apply number formats to the appended data
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=worksheet.max_row - len(csv_data) + 1, max_row=worksheet.max_row, min_col=1, max_col=len(csv_data.columns))):
        for col_idx, cell in enumerate(row):
            cell.number_format = number_formats.get(col_idx, cell.number_format)
    print("Appended data to XLSX file...")
except Exception as e:
    print(f"Error loading XLSX file: {e}")

try:
    # Expand the table to include the new data
    tb = worksheet.tables[table_name]
    current_range = tb.ref
    start_cell, end_cell = current_range.split(':')
    start_row = int(start_cell[1:])
    new_end_cell = f"{colnum_string(len(csv_data.columns))}{worksheet.max_row}"
    tb.ref = f"{start_cell}:{new_end_cell}"
    print("Expanded table...")
except Exception as e:
    print(f"Error expanding table: {e}")

try:
    # Save the modified XLSX file
    workbook.save(xlsx_file_path)
    workbook.close()
    print("Workbook saved...")
    print("Done!")
except Exception as e:
    print(f"Error saving XLSX file: {e}")

